# ======================================================
# Synovia Fusion™ – Express Integration Manager
# ======================================================
# Product       : Synovia Fusion™
# Module        : Express Integration Manager – Avon (AVN)
# Script        : AVN_03_Daily_Road_Jobs_v3_1.py
#
# Version       : 3.1.0
# Author        : Synovia Digital
# Deploy to     : \\PL-AZ-INT-PRD\D_Drive\Applications\
#                   Express_Fusion_2026\Express_Fusion_2026\
#
# Purpose:
#   Read staged AVN orders from [AVN].[vw_Road_Job_Flat],
#   generate a combined Road Jobs XML file for Primeline Express,
#   attach a dispatch Excel workbook, and send an email summary.
#
# Outputs:
#   1. Road_Jobs_ALL_{date}_{time}_{uuid}.xml   → OUTPUT_DIR
#   2. AVN_Dispatch_{date}.xlsx                 → OUTPUT_DIR
#   3. Email with both files attached
#
# Change Log:
#   v3.1.0  10/04/2026  CHG-AVN-001 — Date logic overhaul
#     - CollectionReadyDate / ReadyDate = date.today() (file generation date)
#     - Movement.DepartureDate          = date.today()
#     - Movement.ArrivalDate            = add_working_days(today, 1)
#     - DeliveryRequiredDate            = add_working_days(today, 2)
#     - Weekends + GB/NI/IE bank holidays excluded via holiday calendar
#     - NI postcode fix: BT postcode → country GB (not IE)
#     - INI section fallback: tries [database_server], [database],
#       [Fusion_Express_Production] in order — fixes KeyError crash
#     - DB connection: tries config.config_Include first, falls back to INI
#   v3.0.0  09/04/2026  Combined today+tomorrow; Excel dispatch; email
#   v2.2.0  Earlier     TripStop / ResourceAllocation movement fields
#   v2.1.0  Earlier     vw_Road_Job_Flat schema integration
#
# XML field rules (v3.1.0):
#   DepartureDate              = run_date  (date script runs)
#   ArrivalDate                = run_date + 1 working day
#   CollectionReadyDate        = run_date
#   ReadyDate                  = run_date
#   DeliveryRequiredDate       = run_date + 2 working days
#   DeliveryRequiredTime       = vw_Road_Job_Flat.time_window_start (omit if null)
#   ConsigneeReference         = numeric suffix of trip_name
#   ArrivalLocation            = "Primeline Mallusk"
#   ConsigneeAddressCountryISOCode = GB if postcode starts BT, else delivery_country
#   MessageSenderMovementReference = "AVN_Daily_{date} loading_sheet {datetime}"
# ======================================================

from __future__ import annotations

import configparser
import hashlib
import logging
import os
import re
import shutil
import smtplib
import sys
import traceback
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Optional
from uuid import uuid4

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pyodbc

# ══════════════════════════════════════════════════════════════
# ■  SECTION 1 — CONFIGURATION
# ══════════════════════════════════════════════════════════════

SCRIPT_VERSION = "3.1.0"

PROJECT_ROOT = Path(os.environ.get(
    "FUSION_PROJECT_ROOT",
    r"D:\Applications\Express_Fusion_2026\Express_Fusion_2026"
))

# INI files tried in order
INI_CANDIDATES = [
    Path(r"D:\confguration\fusion_express.ini"),      # deliberate server typo
    Path(r"D:\Configuration\fusion_express.ini"),
    PROJECT_ROOT / "config" / "fusion_express.ini",
]

# INI section names tried in order — prevents KeyError: 'database_server'
INI_SECTION_CANDIDATES = [
    "database_server",
    "database",
    "Fusion_Express_Production",
]

DATABASE_NAME = "Fusion_Express_Production"

OUTPUT_DIR  = Path(r"D:\WIP_out\AVN")
ARCHIVE_DIR = OUTPUT_DIR / "Archive"
LOG_DIR     = Path(r"D:\Logs\AVN")

# XML envelope constants — match live Avon North format exactly
SENDER_ID          = "Avon/Freig/N"
RECIPIENT_ID       = "Primeline Express"
MSG_NAME           = "Road Jobs"
SENDER_NAME        = "Avon Freight Group Ltd (N)"
SENDER_CODE        = "Avon/Freig/N"
SENDER_STREET      = "Unit 29 Heming Road"
SENDER_CITY        = "Redditch"
SENDER_COUNTY      = "Worcestershire"
SENDER_POSTCODE    = "B98ODN"
SENDER_COUNTRY     = "GB"
DEPARTURE_LOCATION = "Redditch"
ARRIVAL_LOCATION   = "Primeline Mallusk"
MOVEMENT_TYPE      = "Collection"
JOB_MOVEMENT_TYPE  = "Delivery"

# ══════════════════════════════════════════════════════════════
# ■  SECTION 2 — LOGGING
# ══════════════════════════════════════════════════════════════

LOG_DIR.mkdir(parents=True, exist_ok=True)
_log_ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
_log_file = LOG_DIR / f"AVN_03_v{SCRIPT_VERSION}_{_log_ts}.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(_log_file, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("AVN_03")

# ══════════════════════════════════════════════════════════════
# ■  SECTION 3 — WORKING DAY CALENDAR  (CHG-AVN-001)
#
#   Confirmed against 1,000+ historical records (AVN mapping workbook):
#     Mon +1 = Tue    Mon +2 = Wed
#     Thu +1 = Fri    Thu +2 = Mon   (skip weekend)
#     Fri +1 = Mon    Fri +2 = Tue   (skip weekend)
#   Bank holidays cause additional skips.
# ══════════════════════════════════════════════════════════════

try:
    import holidays as _holidays_pkg
    _HAS_HOLIDAYS_PKG = True
except ImportError:
    _HAS_HOLIDAYS_PKG = False


def _easter(year: int) -> date:
    """Gregorian Easter Sunday via Anonymous Gregorian algorithm."""
    a = year % 19
    b, c = divmod(year, 100)
    d, e = divmod(b, 4)
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i, k = divmod(c, 4)
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day   = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)


def _first_monday(year: int, month: int) -> date:
    d = date(year, month, 1)
    return d + timedelta(days=(7 - d.weekday()) % 7)


def _last_monday(year: int, month: int) -> date:
    last = date(year, month + 1, 1) - timedelta(days=1) if month < 12 \
           else date(year + 1, 1, 1) - timedelta(days=1)
    return last - timedelta(days=last.weekday())


def _static_holidays(year: int) -> set[date]:
    """
    Static fallback calendar — England/Wales + Northern Ireland + Republic of Ireland.
    Review and update each January.
    """
    easter        = _easter(year)
    good_friday   = easter - timedelta(days=2)
    easter_monday = easter + timedelta(days=1)
    may_day       = _first_monday(year, 5)
    spring_bh     = _last_monday(year, 5)    # England/Wales late May
    summer_bh_gb  = _last_monday(year, 8)    # England/Wales last Aug Mon
    august_roi    = _first_monday(year, 8)   # ROI first Aug Mon
    october_roi   = _last_monday(year, 10)   # ROI last Oct Mon

    return {
        date(year, 1, 1),     # New Year's Day
        good_friday,
        easter_monday,
        may_day,              # Early May bank holiday
        spring_bh,            # Spring bank holiday (Eng/Wales)
        summer_bh_gb,         # Summer bank holiday (Eng/Wales)
        august_roi,           # August bank holiday (ROI)
        october_roi,          # October bank holiday (ROI)
        date(year, 3, 17),    # St Patrick's Day (NI + ROI)
        date(year, 7, 12),    # Battle of the Boyne (NI)
        date(year, 12, 8),    # Immaculate Conception (ROI banks)
        date(year, 12, 25),   # Christmas Day
        date(year, 12, 26),   # Boxing Day / St Stephen's Day
    }


_HOLIDAY_CACHE: dict[int, set[date]] = {}


def _holidays_for(year: int) -> set[date]:
    if year not in _HOLIDAY_CACHE:
        if _HAS_HOLIDAYS_PKG:
            _HOLIDAY_CACHE[year] = (
                set(_holidays_pkg.country_holidays("GB", subdiv="NIR", years=year).keys())
                | set(_holidays_pkg.country_holidays("IE", years=year).keys())
            )
        else:
            log.warning(
                "Package 'holidays' not installed — using static calendar. "
                "Run: pip install holidays"
            )
            _HOLIDAY_CACHE[year] = _static_holidays(year)
    return _HOLIDAY_CACHE[year]


def is_working_day(d: date) -> bool:
    return d.weekday() < 5 and d not in _holidays_for(d.year)


def add_working_days(start: date, n: int) -> date:
    current = start
    count   = 0
    while count < n:
        current += timedelta(days=1)
        if is_working_day(current):
            count += 1
    return current


def fmt_date(d: date) -> str:
    """DD/MM/YYYY — Azyra wire format."""
    return d.strftime("%d/%m/%Y")


def fmt_iso(d: date) -> str:
    """YYYY-MM-DD — filenames."""
    return d.strftime("%Y-%m-%d")


# ══════════════════════════════════════════════════════════════
# ■  SECTION 4 — DATABASE CONNECTION
# ══════════════════════════════════════════════════════════════

def _find_ini() -> Optional[Path]:
    for p in INI_CANDIDATES:
        if p.exists():
            log.info(f"INI: {p}")
            return p
    log.warning(f"INI not found. Tried: {[str(p) for p in INI_CANDIDATES]}")
    return None


def _resolve_ini_section(
    cfg: configparser.ConfigParser,
) -> Optional[configparser.SectionProxy]:
    """
    Try INI_SECTION_CANDIDATES in order.
    Prevents KeyError: 'database_server' when EPOS/Express INI
    uses [database] or another section name.
    """
    for name in INI_SECTION_CANDIDATES:
        if name in cfg:
            log.info(f"INI DB section: [{name}]")
            return cfg[name]
    log.error(
        f"No DB section found. Tried: {INI_SECTION_CANDIDATES}. "
        f"Present: {cfg.sections()}"
    )
    return None


def get_connection() -> pyodbc.Connection:
    """
    Obtain pyodbc connection.
    1. Try Fusion SDK (config.config_Include.connect)
    2. Fall back to direct INI-based connection
    Raises RuntimeError if both fail.
    """
    # ── Attempt 1: Fusion SDK ──────────────────────────────────────────────
    if str(PROJECT_ROOT) not in sys.path:
        sys.path.insert(0, str(PROJECT_ROOT))
    try:
        from config.config_Include import connect  # type: ignore
        conn = connect()
        log.info("DB: connected via config.config_Include")
        return conn
    except Exception as sdk_err:
        log.warning(f"config_Include unavailable ({sdk_err}) — trying INI.")

    # ── Attempt 2: INI ────────────────────────────────────────────────────
    ini_path = _find_ini()
    if ini_path is None:
        raise RuntimeError(
            "Cannot connect: config_Include unavailable and no INI file found. "
            f"Tried: {[str(p) for p in INI_CANDIDATES]}"
        )

    cfg = configparser.ConfigParser()
    cfg.read(str(ini_path), encoding="utf-8")
    sec = _resolve_ini_section(cfg)
    if sec is None:
        raise RuntimeError(
            f"INI {ini_path} has no recognised DB section. "
            f"Sections found: {cfg.sections()}. "
            f"Expected one of: {INI_SECTION_CANDIDATES}"
        )

    driver   = sec.get("driver",   "ODBC Driver 17 for SQL Server").strip("{}")
    server   = sec.get("server",   "")
    database = sec.get("database", DATABASE_NAME)
    user     = sec.get("user",     sec.get("uid", sec.get("username", "")))
    password = sec.get("password", sec.get("pwd", ""))
    encrypt  = sec.get("encrypt",  "yes")
    tsc      = sec.get("trust_server_certificate", "no")

    conn_str = (
        f"DRIVER={{{driver}}};"
        f"SERVER={server};"
        f"DATABASE={database};"
        f"UID={user};"
        f"PWD={password};"
        f"Encrypt={encrypt};"
        f"TrustServerCertificate={tsc};"
    )
    try:
        conn = pyodbc.connect(conn_str, timeout=30)
        log.info(f"DB: connected via INI ({server} / {database})")
        return conn
    except pyodbc.Error as e:
        raise RuntimeError(f"pyodbc connection failed: {e}") from e


# ══════════════════════════════════════════════════════════════
# ■  SECTION 5 — DATA FETCH
# ══════════════════════════════════════════════════════════════

_SQL_ORDERS = """
SELECT
    f.trip_name,
    f.trip_created_ts,
    f.order_identifier,
    f.customer_ref,
    f.customer_name,
    f.delivery_date,
    f.delivery_location_name,
    f.delivery_country,
    f.delivery_address_line1,
    f.delivery_address_line2,
    f.delivery_city,
    f.delivery_state,
    f.delivery_postal_code,
    f.collection_date,
    f.collection_location_name,
    f.collection_country,
    f.collection_address_line1,
    f.collection_city,
    f.collection_state,
    f.goods_description,
    f.total_weight_kg,
    f.pallet_count,
    f.pallet_spaces,
    f.time_window_start,
    f.consignee_instructions
FROM [AVN].[vw_Road_Job_Flat] f
WHERE f.stage_status IN ('STAGED', 'EXPORTED')
ORDER BY f.trip_name, f.order_identifier
"""

_SQL_PIECES = """
SELECT
    h.order_identifier,
    h.barcode,
    h.piece_type
FROM [AVN].[HandlingUnit] h
WHERE EXISTS (
    SELECT 1 FROM [AVN].[vw_Road_Job_Flat] f
    WHERE f.order_identifier = h.order_identifier
      AND f.stage_status IN ('STAGED', 'EXPORTED')
)
ORDER BY h.order_identifier, h.barcode
"""

_SQL_FUTURE = """
SELECT
    COUNT(*)                       AS future_orders,
    SUM(ISNULL(pallet_count, 0))   AS future_pallets
FROM [AVN].[vw_Road_Job_Flat]
WHERE stage_status NOT IN ('STAGED', 'EXPORTED', 'SENT', 'CANCELLED')
"""

_SQL_CHANGES = """
SELECT TOP 10
    order_identifier,
    stage_status,
    CONVERT(VARCHAR, modified_ts, 103) AS change_date
FROM [AVN].[Road_Job_Stage]
WHERE modified_ts >= DATEADD(day, -14, GETUTCDATE())
  AND stage_status IN ('CANCELLED', 'AMENDED')
ORDER BY modified_ts DESC
"""


def fetch_data(
    conn: pyodbc.Connection,
) -> tuple[list[dict], dict[str, list[dict]], dict]:
    cur = conn.cursor()

    # Orders
    cur.execute(_SQL_ORDERS)
    cols   = [c[0] for c in cur.description]
    orders = [dict(zip(cols, r)) for r in cur.fetchall()]

    # Pieces indexed by order_identifier
    cur.execute(_SQL_PIECES)
    pcols  = [c[0] for c in cur.description]
    pieces: dict[str, list[dict]] = defaultdict(list)
    for row in cur.fetchall():
        d = dict(zip(pcols, row))
        pieces[str(d["order_identifier"])].append(d)

    stats: dict = {
        "total_orders":   len(orders),
        "total_pallets":  sum((r.get("pallet_count") or 0) for r in orders),
        "total_weight":   sum(float(r.get("total_weight_kg") or 0) for r in orders),
        "future_orders":  0,
        "future_pallets": 0,
        "changes":        [],
    }

    try:
        cur.execute(_SQL_FUTURE)
        row = cur.fetchone()
        if row:
            stats["future_orders"]  = row[0] or 0
            stats["future_pallets"] = row[1] or 0
    except Exception:
        pass

    try:
        cur.execute(_SQL_CHANGES)
        ccols = [c[0] for c in cur.description]
        stats["changes"] = [dict(zip(ccols, r)) for r in cur.fetchall()]
    except Exception:
        pass

    cur.close()
    log.info(
        f"Fetched: {stats['total_orders']} orders / "
        f"{stats['total_pallets']} pallets / "
        f"{stats['total_weight']:.1f} kg"
    )
    return orders, pieces, stats


# ══════════════════════════════════════════════════════════════
# ■  SECTION 6 — XML HELPERS
# ══════════════════════════════════════════════════════════════

def _xe(v) -> str:
    if v is None:
        return ""
    return (str(v).strip()
            .replace("&", "&amp;")
            .replace("<",  "&lt;")
            .replace(">",  "&gt;")
            .replace('"',  "&quot;"))


def _trip_number(trip_name: str) -> str:
    """'Trip-449451' → '449451'."""
    m = re.search(r"-(\d+)$", trip_name or "")
    return m.group(1) if m else ""


def _consignee_country(row: dict) -> str:
    """BT postcodes → GB (Northern Ireland); otherwise use delivery_country."""
    pc = str(row.get("delivery_postal_code") or "").strip().upper()
    return "GB" if pc.startswith("BT") else _xe(row.get("delivery_country") or "IE")


# ══════════════════════════════════════════════════════════════
# ■  SECTION 7 — XML BUILDER
# ══════════════════════════════════════════════════════════════

def build_job_xml(
    row: dict,
    pieces: list[dict],
    collection_str: str,   # CHG-AVN-001: run_date as DD/MM/YYYY
    delivery_str: str,     # CHG-AVN-001: run_date +2 WD as DD/MM/YYYY
) -> str:
    p = "      "  # indent inside <Jobs>

    order_id = _xe(row.get("order_identifier") or "")
    trip_num = _trip_number(str(row.get("trip_name") or ""))

    coll_name   = _xe(row.get("collection_location_name"))
    coll_street = _xe(row.get("collection_address_line1"))
    coll_city   = _xe(row.get("collection_city"))
    coll_county = _xe(row.get("collection_state"))
    coll_iso    = _xe(row.get("collection_country") or "GB")

    cons_country = _consignee_country(row)
    postcode     = _xe(row.get("delivery_postal_code") or "")

    pallet_qty = row.get("pallet_count") or len(pieces) or 1
    weight     = _xe(row.get("total_weight_kg") or "")
    piece_type = (pieces[0].get("piece_type") if pieces else None) or "Pallets"
    del_time   = _xe(row.get("time_window_start"))
    timed_flag = "Yes" if del_time else "No"

    x  = f"{p}<Job>\n"
    x += f"{p}  <MessageSendersJobReference>{order_id}</MessageSendersJobReference>\n"
    x += f"{p}  <AgentsReference></AgentsReference>\n"
    x += f"{p}  <JobMovementType>{JOB_MOVEMENT_TYPE}</JobMovementType>\n"

    # Sender (always Avon Redditch static values)
    x += f"{p}  <SenderAccountCode>{SENDER_CODE}</SenderAccountCode>\n"
    x += f"{p}  <SenderAddressCode>{SENDER_CODE}</SenderAddressCode>\n"
    x += f"{p}  <SenderName>{SENDER_NAME}</SenderName>\n"
    x += f"{p}  <SenderAddressStreet>{SENDER_STREET}</SenderAddressStreet>\n"
    x += f"{p}  <SenderAddressCity>{SENDER_CITY}</SenderAddressCity>\n"
    x += f"{p}  <SenderAddressCounty>{SENDER_COUNTY}</SenderAddressCounty>\n"
    x += f"{p}  <SenderAddressPostCode>{SENDER_POSTCODE}</SenderAddressPostCode>\n"
    x += f"{p}  <SenderAddressCountryISOCode>{SENDER_COUNTRY}</SenderAddressCountryISOCode>\n"
    x += f"{p}  <SenderJobReference>{order_id}</SenderJobReference>\n"

    # CollectFrom
    if coll_name:
        x += f"{p}  <CollectFromName>{coll_name}</CollectFromName>\n"
    if coll_street:
        x += f"{p}  <CollectFromAddressStreet>{coll_street}</CollectFromAddressStreet>\n"
    if coll_city:
        x += f"{p}  <CollectFromAddressCity>{coll_city}</CollectFromAddressCity>\n"
    if coll_county:
        x += f"{p}  <CollectFromAddressCounty>{coll_county}</CollectFromAddressCounty>\n"
    x += f"{p}  <CollectFromAddressCountryISOCode>{coll_iso}</CollectFromAddressCountryISOCode>\n"

    # CHG-AVN-001: collection date = run_date
    x += f"{p}  <CollectionReadyDate>{collection_str}</CollectionReadyDate>\n"
    x += f"{p}  <ReadyDate>{collection_str}</ReadyDate>\n"

    # Consignee
    x += f"{p}  <ConsigneeName>{_xe(row.get('delivery_location_name'))}</ConsigneeName>\n"
    x += f"{p}  <ConsigneeReference>{trip_num}</ConsigneeReference>\n"
    x += f"{p}  <ConsigneeAddressStreet>{_xe(row.get('delivery_address_line1'))}</ConsigneeAddressStreet>\n"
    if row.get("delivery_city"):
        x += f"{p}  <ConsigneeAddressCity>{_xe(row['delivery_city'])}</ConsigneeAddressCity>\n"
    if row.get("delivery_state"):
        x += f"{p}  <ConsigneeAddressCounty>{_xe(row['delivery_state'])}</ConsigneeAddressCounty>\n"
    x += f"{p}  <ConsigneeAddressPostCode>{postcode}</ConsigneeAddressPostCode>\n"
    x += f"{p}  <ConsigneeAddressCountryISOCode>{cons_country}</ConsigneeAddressCountryISOCode>\n"
    if row.get("consignee_instructions"):
        x += f"{p}  <ConsigneeInstructions>{_xe(row['consignee_instructions'])}</ConsigneeInstructions>\n"

    # CHG-AVN-001: delivery date = run_date + 2 working days
    x += f"{p}  <DeliveryRequiredDate>{delivery_str}</DeliveryRequiredDate>\n"
    if del_time:
        x += f"{p}  <DeliveryRequiredTime>{del_time}</DeliveryRequiredTime>\n"

    # Goods
    x += f"{p}  <GoodsDescription>{_xe(row.get('goods_description'))}</GoodsDescription>\n"
    x += f"{p}  <GoodsWeightGrossKgs>{weight}</GoodsWeightGrossKgs>\n"
    x += f"{p}  <GoodsChargeableKgs>{weight}</GoodsChargeableKgs>\n"
    x += f"{p}  <GoodsPiecesQuantity1>{pallet_qty}</GoodsPiecesQuantity1>\n"
    x += f"{p}  <GoodsPiecesDescription1>{_xe(piece_type)}</GoodsPiecesDescription1>\n"

    # Piece barcodes
    if pieces:
        x += f"{p}  <Pieces>\n"
        for pc in pieces:
            x += f"{p}    <Piece>\n"
            x += f"{p}      <SenderPieceID>{_xe(pc.get('barcode'))}</SenderPieceID>\n"
            x += f"{p}      <PieceType>{_xe(pc.get('piece_type') or 'Pallets')}</PieceType>\n"
            x += f"{p}    </Piece>\n"
        x += f"{p}  </Pieces>\n"

    # Ancillary
    x += f"{p}  <SpecialDate01>{collection_str}</SpecialDate01>\n"
    x += f"{p}  <SpecialText01>{_xe(row.get('customer_ref'))}</SpecialText01>\n"
    x += f"{p}  <SpecialText02></SpecialText02>\n"
    x += f"{p}  <SpecialText03></SpecialText03>\n"
    x += f"{p}  <TimedDelivery>{timed_flag}</TimedDelivery>\n"
    x += f"{p}</Job>\n"
    return x


def build_xml(
    orders: list[dict],
    pieces_map: dict[str, list[dict]],
    run_date: date,
) -> tuple[str, str]:
    """
    Build the complete Road Jobs XML.
    Returns (xml_string, filename).
    All dates derived from run_date (CHG-AVN-001).
    """
    arrival_date  = add_working_days(run_date, 1)
    delivery_date = add_working_days(run_date, 2)

    collection_str = fmt_date(run_date)
    arrival_str    = fmt_date(arrival_date)
    delivery_str   = fmt_date(delivery_date)

    log.info(
        f"CHG-AVN-001: Collection={collection_str}  "
        f"Arrival={arrival_str}  Delivery={delivery_str}"
    )

    now_utc  = datetime.now(timezone.utc)
    msg_id   = hashlib.md5(str(uuid4()).encode()).hexdigest()[:12].upper()
    msg_date = now_utc.strftime("%d/%m/%Y")
    msg_time = now_utc.strftime("%I:%M:%S %p")
    now_tag  = now_utc.strftime("%Y%m%d_%H%M%S")
    uid8     = str(uuid4()).replace("-", "")[:8].upper()

    mov_ref = (
        f"AVN_Daily_{fmt_iso(run_date)} "
        f"loading_sheet {now_utc.strftime('%d/%m/%Y %I:%M:%S %p')}"
    )

    xml  = '<?xml version="1.0" encoding="UTF-8"?>\n'
    xml += "<RoadJobs>\n"
    xml += "  <Message>\n"
    xml += f"    <MessageName>{MSG_NAME}</MessageName>\n"
    xml += f"    <MessageID>{msg_id}</MessageID>\n"
    xml += f"    <MessageSenderID>{SENDER_ID}</MessageSenderID>\n"
    xml += f"    <MessageRecipientID>{RECIPIENT_ID}</MessageRecipientID>\n"
    xml += f"    <MessageCreatedDate>{msg_date}</MessageCreatedDate>\n"
    xml += f"    <MessageCreatedTime>{msg_time}</MessageCreatedTime>\n"
    xml += "    <Movement>\n"
    xml += f"      <MessageSenderMovementReference>{_xe(mov_ref)}</MessageSenderMovementReference>\n"
    xml += f"      <CarrierName>{SENDER_NAME}</CarrierName>\n"
    xml += f"      <MovementType>{MOVEMENT_TYPE}</MovementType>\n"
    xml += f"      <DepartureDate>{collection_str}</DepartureDate>\n"
    xml += f"      <DepartureLocation>{DEPARTURE_LOCATION}</DepartureLocation>\n"
    xml += f"      <ArrivalDate>{arrival_str}</ArrivalDate>\n"
    xml += f"      <ArrivalLocation>{ARRIVAL_LOCATION}</ArrivalLocation>\n"
    xml += "      <Jobs>\n"

    for row in orders:
        oid = str(row.get("order_identifier") or "")
        pcs = pieces_map.get(oid, [])
        xml += build_job_xml(row, pcs, collection_str, delivery_str)

    xml += "      </Jobs>\n"
    xml += "    </Movement>\n"
    xml += "  </Message>\n"
    xml += "</RoadJobs>\n"

    filename = f"Road_Jobs_ALL_{fmt_iso(run_date)}_{now_tag}_{uid8}.xml"
    return xml, filename


# ══════════════════════════════════════════════════════════════
# ■  SECTION 8 — DISPATCH EXCEL WORKBOOK
# ══════════════════════════════════════════════════════════════

def _thin_border(col: str = "CCCCCC") -> Border:
    s = Side(style="thin", color=col)
    return Border(left=s, right=s, top=s, bottom=s)


def _solid(hex_: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_)


def build_dispatch_workbook(
    orders: list[dict],
    pieces_map: dict[str, list[dict]],
    run_date: date,
    delivery_date: date,
    output_path: Path,
) -> None:
    wb = openpyxl.Workbook()
    bd = _thin_border()

    # ──────────────────────────────────────────────────────────
    # Tab 1 — Dispatch Summary
    # ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Dispatch Summary"

    ws.merge_cells("A1:J1")
    ws["A1"].value = (
        f"AVN Road Jobs Dispatch  |  Collection: {fmt_date(run_date)}"
        f"  |  Delivery: {fmt_date(delivery_date)}"
        f"  |  Generated: {datetime.now():%d/%m/%Y %H:%M}"
    )
    ws["A1"].font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    ws["A1"].fill      = _solid("0D1B2A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    dispatch_headers = [
        "Trip", "Order Ref", "Consignee", "Street",
        "City", "Postcode", "Country", "Weight (kg)", "Pallets", "Timed?"
    ]
    for ci, h in enumerate(dispatch_headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        c.fill      = _solid("1B3A5C")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = bd
    ws.row_dimensions[2].height = 18

    alt_fill = _solid("F5F8FF")
    for ri, row in enumerate(orders, 3):
        is_alt = ri % 2 == 0
        vals = [
            row.get("trip_name"),
            row.get("order_identifier"),
            row.get("delivery_location_name"),
            row.get("delivery_address_line1"),
            row.get("delivery_city"),
            row.get("delivery_postal_code"),
            _consignee_country(row),
            row.get("total_weight_kg"),
            row.get("pallet_count"),
            "Yes" if row.get("time_window_start") else "No",
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font   = Font(name="Arial", size=9)
            c.border = bd
            if is_alt:
                c.fill = alt_fill

    col_widths = [18, 14, 28, 32, 16, 12, 10, 13, 10, 9]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes below header
    ws.freeze_panes = "A3"

    # ──────────────────────────────────────────────────────────
    # Tab 2 — Field Mapping (reference)
    # ──────────────────────────────────────────────────────────
    wm = wb.create_sheet("Field Mapping")

    map_headers = ["XML Field", "Source", "Calculation Rule v3.1.0", "Changed?"]
    for ci, h in enumerate(map_headers, 1):
        c = wm.cell(row=1, column=ci, value=h)
        c.font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        c.fill      = _solid("1B3A5C")
        c.border    = bd
        c.alignment = Alignment(horizontal="center")

    chg_fill = _solid("FFF3E0")
    chg_rows = {
        "DepartureDate", "ArrivalDate",
        "CollectionReadyDate", "ReadyDate", "DeliveryRequiredDate",
    }

    arrival_str_ex  = fmt_date(add_working_days(run_date, 1))
    delivery_str_ex = fmt_date(delivery_date)

    mapping_data = [
        ("DepartureDate",               "Computed",          f"date.today() = {fmt_date(run_date)}",              "YES"),
        ("ArrivalDate",                 "Computed",          f"run_date +1 WD = {arrival_str_ex}",                "YES"),
        ("CollectionReadyDate",         "Computed",          f"date.today() = {fmt_date(run_date)}",              "YES"),
        ("ReadyDate",                   "Computed",          f"date.today() = {fmt_date(run_date)}",              "YES"),
        ("DeliveryRequiredDate",        "Computed",          f"run_date +2 WD = {delivery_str_ex}",               "YES"),
        ("DeliveryRequiredTime",        "vw_Road_Job_Flat",  "time_window_start — omit element if NULL",          "No"),
        ("MessageSenderMovementRef",    "Computed",          "AVN_Daily_{run_date} loading_sheet {datetime}",     "No"),
        ("ConsigneeName",               "vw_Road_Job_Flat",  "delivery_location_name",                            "No"),
        ("ConsigneeAddressStreet",      "vw_Road_Job_Flat",  "delivery_address_line1",                            "No"),
        ("ConsigneeAddressCity",        "vw_Road_Job_Flat",  "delivery_city — omit if NULL",                      "No"),
        ("ConsigneeAddressCounty",      "vw_Road_Job_Flat",  "delivery_state — omit if NULL",                     "No"),
        ("ConsigneeAddressPostCode",    "vw_Road_Job_Flat",  "delivery_postal_code",                              "No"),
        ("ConsigneeCountryISOCode",     "Computed",          "GB if postcode starts BT, else delivery_country",   "No"),
        ("ConsigneeReference",          "Computed",          "Numeric suffix of trip_name",                       "No"),
        ("GoodsPiecesQuantity1",        "vw_Road_Job_Flat",  "pallet_count",                                      "No"),
        ("GoodsDescription",            "vw_Road_Job_Flat",  "goods_description",                                 "No"),
        ("GoodsWeightGrossKgs",         "vw_Road_Job_Flat",  "total_weight_kg",                                   "No"),
        ("GoodsChargeableKgs",          "vw_Road_Job_Flat",  "total_weight_kg (same source)",                     "No"),
        ("SenderPieceID",               "AVN.HandlingUnit",  "barcode",                                           "No"),
        ("PieceType",                   "AVN.HandlingUnit",  "piece_type",                                        "No"),
        ("SenderName",                  "Static",            f"'{SENDER_NAME}'",                                  "No"),
        ("SenderAddressStreet",         "Static",            f"'{SENDER_STREET}'",                                "No"),
        ("SenderAddressCity",           "Static",            f"'{SENDER_CITY}'",                                  "No"),
        ("SenderAddressPostCode",       "Static",            f"'{SENDER_POSTCODE}'",                              "No"),
        ("SenderAddressCountryISOCode", "Static",            "GB",                                                "No"),
        ("ArrivalLocation",             "Static",            f"'{ARRIVAL_LOCATION}'",                             "No"),
        ("DepartureLocation",           "Static",            f"'{DEPARTURE_LOCATION}'",                           "No"),
    ]

    for ri, (field, src, rule, chg) in enumerate(mapping_data, 2):
        row_fill = chg_fill if field in chg_rows else _solid("FFFFFF")
        for ci, v in enumerate([field, src, rule, chg], 1):
            c = wm.cell(row=ri, column=ci, value=v)
            c.font      = Font(name="Arial", size=9,
                               bold=(chg == "YES"),
                               color=("E65100" if chg == "YES" else "000000"))
            c.fill      = row_fill
            c.border    = bd
            c.alignment = Alignment(wrap_text=True, vertical="top")

    wm.column_dimensions["A"].width = 32
    wm.column_dimensions["B"].width = 20
    wm.column_dimensions["C"].width = 44
    wm.column_dimensions["D"].width = 12

    # ──────────────────────────────────────────────────────────
    # Tab 3 — Feedback form
    # ──────────────────────────────────────────────────────────
    wf = wb.create_sheet("Feedback")

    wf.merge_cells("A1:E1")
    wf["A1"].value = (
        "AVN Data Quality Feedback — complete and return to aidan.harrington@synoviadigital.com"
    )
    wf["A1"].font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    wf["A1"].fill      = _solid("0D1B2A")
    wf["A1"].alignment = Alignment(horizontal="center", vertical="center")
    wf.row_dimensions[1].height = 22

    fb_headers = ["Order Ref", "Issue Type", "Field Affected", "Correct Value", "Notes"]
    for ci, h in enumerate(fb_headers, 1):
        c = wf.cell(row=2, column=ci, value=h)
        c.font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        c.fill      = _solid("1B3A5C")
        c.border    = bd
        c.alignment = Alignment(horizontal="center")

    for ri in range(3, 25):
        for ci in range(1, 6):
            c = wf.cell(row=ri, column=ci)
            c.border = bd
            c.font   = Font(name="Arial", size=9)

    for w, col in zip([16, 22, 22, 28, 32], "ABCDE"):
        wf.column_dimensions[col].width = w

    wb.save(str(output_path))
    log.info(f"Dispatch workbook saved: {output_path}")


# ══════════════════════════════════════════════════════════════
# ■  SECTION 9 — EMAIL
# ══════════════════════════════════════════════════════════════

def _email_html(
    stats: dict,
    run_date: date,
    delivery_date: date,
    xml_filename: str,
    xl_filename: str,
) -> str:
    chg_rows_html = ""
    for ch in stats.get("changes", []):
        chg_rows_html += (
            f"<tr style='font-size:12px'>"
            f"<td style='padding:5px 10px'>{ch.get('order_identifier','')}</td>"
            f"<td style='padding:5px 10px'>{ch.get('stage_status','')}</td>"
            f"<td style='padding:5px 10px'>{ch.get('change_date','')}</td>"
            f"</tr>"
        )

    chg_section = ""
    if chg_rows_html:
        chg_section = f"""
        <h3 style="color:#B71C1C;margin-top:22px">
          &#9888; Recent Changes / Cancellations (last 14 days)
        </h3>
        <table style="border-collapse:collapse;font-size:12px;width:auto">
          <tr style="background:#1B3A5C;color:#FFF">
            <th style="padding:6px 12px;text-align:left">Order Ref</th>
            <th style="padding:6px 12px;text-align:left">Status</th>
            <th style="padding:6px 12px;text-align:left">Date</th>
          </tr>
          {chg_rows_html}
        </table>"""

    return f"""<html>
<body style="font-family:Arial,sans-serif;font-size:13px;color:#1A1A2E;margin:0;padding:20px">

<div style="background:#0D1B2A;color:#FFF;padding:18px 24px;border-radius:6px;margin-bottom:22px">
  <h2 style="margin:0 0 6px 0;font-size:18px">Synovia Fusion&#8482; &#8212; AVN Road Jobs</h2>
  <p style="margin:0;font-size:11px;color:#A0B4CC">
    AVN_03_Daily_Road_Jobs_v{SCRIPT_VERSION} &#124; CHG-AVN-001 date rules active
  </p>
</div>

<table style="border-collapse:collapse;margin-bottom:18px">
  <tr><td style="padding:4px 20px 4px 0;font-weight:bold">Collection Date</td>
      <td>{fmt_date(run_date)} ({run_date.strftime('%A')})</td></tr>
  <tr><td style="padding:4px 20px 4px 0;font-weight:bold">Arrival at Mallusk</td>
      <td>{fmt_date(add_working_days(run_date, 1))}</td></tr>
  <tr><td style="padding:4px 20px 4px 0;font-weight:bold">Delivery Date</td>
      <td>{fmt_date(delivery_date)} ({delivery_date.strftime('%A')})</td></tr>
  <tr><td style="padding:4px 20px 4px 0;font-weight:bold">Total Orders</td>
      <td>{stats['total_orders']}</td></tr>
  <tr><td style="padding:4px 20px 4px 0;font-weight:bold">Total Pallets</td>
      <td>{stats['total_pallets']}</td></tr>
  <tr><td style="padding:4px 20px 4px 0;font-weight:bold">Total Weight</td>
      <td>{stats['total_weight']:.1f} kg</td></tr>
  <tr><td style="padding:4px 20px 4px 0;font-weight:bold">Future Pipeline</td>
      <td>{stats['future_orders']} orders / {stats['future_pallets']} pallets
          (not in this file)</td></tr>
</table>

<p style="margin-bottom:6px"><strong>Attachments:</strong></p>
<ol style="margin:0;padding-left:20px;line-height:1.8">
  <li><b>{xml_filename}</b> &#8212; Road Jobs XML for Pure / Primeline upload</li>
  <li><b>{xl_filename}</b> &#8212; Dispatch workbook
    (Dispatch Summary / Field Mapping / Feedback tabs)</li>
</ol>

{chg_section}

<p style="margin-top:24px;font-size:11px;color:#888">
  Synovia Fusion&#8482; &#124; Avon Freight North (AVN) &#124; Primeline Express Integration
</p>
</body>
</html>"""


def send_email(
    run_date: date,
    delivery_date: date,
    stats: dict,
    xml_path: Path,
    xl_path: Path,
) -> None:
    ini_path = _find_ini()
    if ini_path is None:
        log.info("Email: no INI — skipped.")
        return

    cfg = configparser.ConfigParser()
    cfg.read(str(ini_path), encoding="utf-8")

    if "email" not in cfg:
        log.info("Email: no [email] section in INI — skipped.")
        return

    sec        = cfg["email"]
    smtp_host  = sec.get("smtp_host",     "smtp.office365.com")
    smtp_port  = int(sec.get("smtp_port", 587))
    from_addr  = sec.get("from_address",  "fusion@synoviadigital.com")
    recipients = [r.strip() for r in sec.get("to_addresses", "").split(",") if r.strip()]

    if not recipients:
        log.info("Email: no recipients configured — skipped.")
        return

    subject = (
        f"AVN Road Jobs \u2014 Collection {fmt_date(run_date)} "
        f"/ Delivery {fmt_date(delivery_date)} "
        f"\u2014 {stats['total_orders']} orders / {stats['total_pallets']} pallets"
    )

    msg            = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"]    = from_addr
    msg["To"]      = ", ".join(recipients)
    msg.attach(MIMEText(
        _email_html(stats, run_date, delivery_date, xml_path.name, xl_path.name),
        "html"
    ))

    for p in [xml_path, xl_path]:
        if p and p.exists():
            with open(p, "rb") as f:
                part = MIMEApplication(f.read(), Name=p.name)
            part["Content-Disposition"] = f'attachment; filename="{p.name}"'
            msg.attach(part)

    try:
        with smtplib.SMTP(smtp_host, smtp_port) as smtp:
            smtp.ehlo()
            smtp.starttls()
            pw = sec.get("password", "")
            un = sec.get("username", from_addr)
            if pw:
                smtp.login(un, pw)
            smtp.sendmail(from_addr, recipients, msg.as_string())
        log.info(f"Email sent to: {', '.join(recipients)}")
    except Exception as e:
        log.error(f"Email failed (outputs already written to disk): {e}")


# ══════════════════════════════════════════════════════════════
# ■  SECTION 10 — MAIN
# ══════════════════════════════════════════════════════════════

def main() -> None:
    run_date      = date.today()
    arrival_date  = add_working_days(run_date, 1)
    delivery_date = add_working_days(run_date, 2)

    log.info("=" * 70)
    log.info(f"Synovia Fusion\u2122  |  AVN_03_Daily_Road_Jobs  v{SCRIPT_VERSION}")
    log.info(f"Run date      : {fmt_date(run_date)}  ({run_date.strftime('%A')})")
    log.info(f"Arrival       : {fmt_date(arrival_date)}  (run_date +1 WD)")
    log.info(f"Delivery      : {fmt_date(delivery_date)}  (run_date +2 WD) \u2190 CHG-AVN-001")
    log.info(f"Holiday pkg   : {'holidays (dynamic)' if _HAS_HOLIDAYS_PKG else 'static fallback'}")
    log.info(f"Output dir    : {OUTPUT_DIR}")
    log.info("=" * 70)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)

    # Connect
    conn = get_connection()

    # Fetch
    orders, pieces, stats = fetch_data(conn)
    conn.close()

    if not orders:
        log.warning("No STAGED/EXPORTED orders found \u2014 nothing to generate.")
        return

    # XML
    xml_content, xml_filename = build_xml(orders, pieces, run_date)
    xml_path = OUTPUT_DIR / xml_filename
    xml_path.write_text(xml_content, encoding="utf-8")
    log.info(f"XML  \u2192 {xml_path}  ({xml_path.stat().st_size:,} bytes)")

    # Dispatch workbook
    xl_filename = f"AVN_Dispatch_{fmt_iso(run_date)}.xlsx"
    xl_path     = OUTPUT_DIR / xl_filename
    build_dispatch_workbook(orders, pieces, run_date, delivery_date, xl_path)

    # Archive copies
    for src in [xml_path, xl_path]:
        shutil.copy2(src, ARCHIVE_DIR / src.name)
    log.info(f"Archived to {ARCHIVE_DIR}")

    # Email
    send_email(run_date, delivery_date, stats, xml_path, xl_path)

    log.info("AVN_03 complete.")
    log.info("=" * 70)


if __name__ == "__main__":
    try:
        main()
    except Exception:
        log.critical("Unhandled exception:\n" + traceback.format_exc())
        sys.exit(1)
